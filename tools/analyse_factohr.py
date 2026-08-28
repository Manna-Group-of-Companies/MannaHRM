"""Answer the migration questions from the Factor HR exports.

Counts and coverage, never a person's details. The questions it exists to
answer, in order of how much they decide:

  1. How many people, per company, and how many are still employed?
  2. **How many have a biometric machine code?** Without one, that person's
     fingerprint punches have nowhere to land, and they look absent every day.
  3. What shifts and week-offs are actually in use?
  4. What terminals do the punch reports name — i.e. how many devices exist?

    python tools/analyse_factohr.py
"""

import os
import sys
from collections import Counter

try:
	from openpyxl import load_workbook
except ImportError:
	sys.exit("Needs openpyxl: python -m pip install openpyxl")

FOLDER = sys.argv[1] if len(sys.argv) > 1 else "data/factohr"


def read_rows(path, sheet=None, limit=2000):
	book = load_workbook(path, read_only=True, data_only=True)
	ws = book[sheet] if sheet else book.worksheets[0]
	rows = [list(r) for r in ws.iter_rows(max_row=limit, values_only=True)]
	book.close()
	return rows


def find_header(rows, must_contain, limit=15):
	"""The row holding real column labels, identified by a column we expect.

	Factor HR stacks a banner, a group-header row and the labels. Searching for
	a known label beats counting rows, because the banner height differs per
	report.
	"""
	want = must_contain.lower()
	for i, row in enumerate(rows[:limit]):
		for cell in row:
			if cell and str(cell).strip().lower() == want:
				return i
	return None


def columns_of(row):
	return [str(c).strip() if c is not None else "" for c in row]


def col(headers, *names):
	"""Index of the first header matching any of `names`, else None."""
	lowered = [h.lower() for h in headers]
	for name in names:
		n = name.lower()
		for i, h in enumerate(lowered):
			if h == n:
				return i
	# Fall back to a prefix match — Factor HR truncates some labels.
	for name in names:
		n = name.lower()
		for i, h in enumerate(lowered):
			if h.startswith(n):
				return i
	return None


def get(row, index):
	if index is None or index >= len(row):
		return ""
	value = row[index]
	return "" if value is None else str(value).strip()


def show(title, counter, total=None, limit=25):
	print("\n  {0}".format(title))
	for key, n in counter.most_common(limit):
		share = ""
		if total:
			share = "  ({0:.0f}%)".format(100.0 * n / total)
		print("    {0:<42} {1:>5}{2}".format((key or "(blank)")[:42], n, share))
	if len(counter) > limit:
		print("    ... and {0} more".format(len(counter) - limit))


# ----------------------------------------------------------------- master ---


def analyse_employees(path):
	print("=" * 74)
	print("EMPLOYEE MASTER  —  {0}".format(os.path.basename(path)))
	print("=" * 74)

	rows = read_rows(path, limit=3000)
	index = find_header(rows, "Emp Code")
	if index is None:
		print("  could not find an 'Emp Code' column")
		return

	headers = columns_of(rows[index])
	data = [r for r in rows[index + 1 :] if any(c not in (None, "") for c in r)]

	i_code = col(headers, "Emp Code")
	i_company = col(headers, "Company Name")
	i_status = col(headers, "Status")
	i_machine = col(headers, "Machine Code")
	i_shift = col(headers, "Working Shift")
	i_weekoff = col(headers, "Week-off", "Week off")
	i_dept = col(headers, "Department")
	i_mgr = col(headers, "Reporting Manager")
	i_type = col(headers, "Employment Type")
	i_paygroup = col(headers, "Payroll Group")
	i_leaving = col(headers, "Leaving Date")

	# A row is a person only if it carries a code; Factor HR pads the sheet.
	people = [r for r in data if get(r, i_code)]
	print("\n  rows with an employee code : {0}".format(len(people)))

	active = [r for r in people if get(r, i_status).lower() == "active"]
	print("  active                     : {0}".format(len(active)))
	print("  not active                 : {0}".format(len(people) - len(active)))

	show("Company", Counter(get(r, i_company) for r in people), len(people))
	show("Status", Counter(get(r, i_status) for r in people), len(people))

	# ---- the one that decides the biometric leg ----
	print("\n  " + "-" * 70)
	print("  BIOMETRIC MACHINE CODE — active employees only")
	print("  " + "-" * 70)

	with_code = [r for r in active if get(r, i_machine)]
	without = [r for r in active if not get(r, i_machine)]
	print("    have a Machine Code      : {0} of {1}".format(len(with_code), len(active)))
	print("    MISSING a Machine Code   : {0}".format(len(without)))

	if without:
		show("    missing, by company", Counter(get(r, i_company) for r in without))

	codes = [get(r, i_machine) for r in with_code]
	dupes = Counter(c for c in codes if c)
	clashing = {c: n for c, n in dupes.items() if n > 1}
	if clashing:
		print("\n    *** {0} machine code(s) used by more than one active person:".format(len(clashing)))
		for code, n in sorted(clashing.items(), key=lambda kv: -kv[1])[:15]:
			owners = [get(r, i_company) for r in with_code if get(r, i_machine) == code]
			print("        code {0:<8} used {1}x  {2}".format(code, n, ", ".join(sorted(set(owners)))[:44]))
		print("        A shared code means those punches cannot be told apart.")
	else:
		print("\n    every machine code is unique among active staff")

	show("Working Shift (active)", Counter(get(r, i_shift) for r in active), len(active))
	show("Week-off (active)", Counter(get(r, i_weekoff) for r in active), len(active))
	show("Employment Type (active)", Counter(get(r, i_type) for r in active), len(active))
	show("Payroll Group (active)", Counter(get(r, i_paygroup) for r in active), len(active))
	show("Department (active)", Counter(get(r, i_dept) for r in active), len(active), limit=15)

	no_mgr = [r for r in active if not get(r, i_mgr)]
	print("\n  active with no Reporting Manager : {0}".format(len(no_mgr)))

	left_no_date = [
		r for r in people
		if get(r, i_status).lower() != "active" and not get(r, i_leaving)
	]
	print("  inactive with no Leaving Date    : {0}".format(len(left_no_date)))


# ------------------------------------------------------------- terminals ---


def analyse_punches(path):
	print("\n" + "=" * 74)
	print("PUNCH SOURCE  —  {0}".format(os.path.basename(path)))
	print("=" * 74)

	rows = read_rows(path, limit=5000)
	index = find_header(rows, "Emp Code")
	if index is None:
		print("  could not find an 'Emp Code' column")
		return

	headers = columns_of(rows[index])
	data = rows[index + 1 :]

	i_code = col(headers, "Emp Code")
	i_term = col(headers, "Terminal")
	i_loc = col(headers, "Location")
	i_info = col(headers, "Punch Info")
	i_selfie = col(headers, "Selfie Image")

	punches = [r for r in data if get(r, i_code)]
	print("\n  punch rows        : {0}".format(len(punches)))
	print("  distinct people   : {0}".format(len(set(get(r, i_code) for r in punches))))

	show("Terminal", Counter(get(r, i_term) for r in punches), len(punches))

	with_loc = sum(1 for r in punches if get(r, i_loc))
	with_info = sum(1 for r in punches if get(r, i_info))
	with_selfie = sum(1 for r in punches if get(r, i_selfie))
	print("\n  punches carrying a Location  : {0}".format(with_loc))
	print("  punches carrying Punch Info  : {0}".format(with_info))
	print("  punches carrying a Selfie    : {0}".format(with_selfie))

	# `Punch Info` is where the mobile app records GPS quality. Worth seeing the
	# shape of, because it decides whether a geofence can be reconstructed.
	samples = [get(r, i_info) for r in punches if get(r, i_info)][:3]
	for s in samples:
		print("    punch info sample : {0}".format(s[:88]))


def main():
	files = sorted(os.listdir(FOLDER))
	for name in files:
		low = name.lower()
		path = os.path.join(FOLDER, name)
		if not low.endswith(".xlsx"):
			continue
		if "employee detail" in low:
			analyse_employees(path)
	for name in files:
		low = name.lower()
		path = os.path.join(FOLDER, name)
		if low.endswith(".xlsx") and "attendance report" in low:
			analyse_punches(path)


if __name__ == "__main__":
	main()
