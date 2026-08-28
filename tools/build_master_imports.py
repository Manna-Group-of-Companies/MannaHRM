"""Generate the ERPNext Data Import files for the HR masters.

    python tools/build_master_imports.py

Writes into `data/out/` (gitignored):

    01-holiday-list.xlsx     Holiday List + its holidays, parent and child
    02-designation.xlsx      the designations the employee import needs

The holiday year follows the site's fiscal year, **1 Apr 2026 - 31 Mar 2027**,
read from ERPNext rather than assumed.

## The dates, and which of them are guesses

Four statutory holidays are fixed by law and by date. The rest move, and their
2026-27 dates were looked up rather than calculated:

  Good Friday  3 Apr 2026   moves with Easter
  Vishu       15 Apr 2026   Malayalam solar new year
  Onam        25-26 Aug 2026  Uthradam and Thiruvonam
  Christmas   25 Dec 2026   fixed

**Check Onam before importing.** Thiruvonam 2026 is 26 August, and whether the
factory closes on Uthradam the day before is a Manna decision, not a calendar
fact. Getting it wrong marks everybody absent on a day they were right to be at
home, and that flows into pay.
"""

import os
import sys
from datetime import date, timedelta

try:
	from openpyxl import Workbook
except ImportError:
	sys.exit("Needs openpyxl: python -m pip install openpyxl")

OUT_DIR = "data/out"

# Read from the site: the only Fiscal Year on it is 2026-04-01 to 2027-03-31.
YEAR_START = date(2026, 4, 1)
YEAR_END = date(2027, 3, 31)

LIST_NAME = "Manna Holidays 2026-27"
OPTIONAL_LIST_NAME = "Manna Optional Holidays 2026-27"

# Sunday. Confirmed as the weekly off for all 160 active employees, with no
# exceptions anywhere in the Factor HR master.
WEEKLY_OFF_WEEKDAY = 6  # Monday=0 ... Sunday=6

STATUTORY = [
	(date(2027, 1, 26), "Republic Day"),
	(date(2026, 5, 1), "May Day"),
	(date(2026, 8, 15), "Independence Day"),
	(date(2026, 10, 2), "Gandhi Jayanti"),
]

# Movable, except Christmas. Dates looked up for 2026-27 - see module docstring.
OPTIONAL = [
	(date(2026, 4, 3), "Good Friday"),
	(date(2026, 4, 15), "Vishu"),
	(date(2026, 8, 25), "Onam - Uthradam"),
	(date(2026, 8, 26), "Onam - Thiruvonam"),
	(date(2026, 12, 25), "Christmas"),
]


def sundays(start, end):
	d = start
	# Walk to the first Sunday rather than testing every day from April.
	d += timedelta(days=(WEEKLY_OFF_WEEKDAY - d.weekday()) % 7)
	while d <= end:
		yield d
		d += timedelta(days=7)


def build_holiday_rows(include_optional):
	"""(date, description, is_weekly_off) sorted, with clashes resolved.

	A holiday landing on a Sunday is kept **once**, as the named holiday rather
	than as the weekly off. Two rows for one date makes ERPNext count the day
	twice, and the second one is what a payroll query trips over.
	"""
	named = {d: label for d, label in STATUTORY}
	if include_optional:
		named.update({d: label for d, label in OPTIONAL})

	rows = []
	for d in sundays(YEAR_START, YEAR_END):
		if d in named:
			continue
		rows.append((d, "Sunday", 1))
	for d, label in named.items():
		if not (YEAR_START <= d <= YEAR_END):
			raise SystemExit("{0} ({1}) falls outside the holiday year".format(label, d))
		rows.append((d, label, 0))

	rows.sort(key=lambda r: r[0])
	return rows


# The exact header row ERPNext produces for this doctype, taken from its own
# Download Template rather than guessed. Two earlier guesses failed: the child
# column is "<Field Label> (<Child Table Label>)" — `Date (Holidays)` — and the
# child date field is labelled **Date**, not Holiday Date.
HOLIDAY_HEADERS = [
	"Holiday List Name", "From Date", "To Date", "Total Holidays",
	"Weekly Off", "Is Half Day", "Country", "Subdivision", "Color",
	"ID (Holidays)", "Date (Holidays)", "Description (Holidays)",
	"Is Half Day (Holidays)", "Weekly Off (Holidays)",
]


def holiday_sheet_rows(list_name, rows):
	"""Parent columns on the first row only; blank on every continuation.

	That blankness is how the importer knows a row belongs to the document
	above it rather than starting a new one. `Total Holidays` is left empty
	throughout — ERPNext computes it, and a supplied value is either ignored or
	wrong.
	"""
	out = []
	for i, (d, label, weekly) in enumerate(rows):
		if i == 0:
			out.append([list_name, YEAR_START.isoformat(), YEAR_END.isoformat(), "",
			            "Sunday", 0, "India", "", "",
			            "", d.isoformat(), label, 0, weekly])
		else:
			out.append(["", "", "", "", "", "", "", "", "",
			            "", d.isoformat(), label, 0, weekly])
	return out


def write_holiday_list(path):
	wb = Workbook()
	ws = wb.active
	ws.title = "Holiday List"
	ws.append(HOLIDAY_HEADERS)

	rows = build_holiday_rows(include_optional=True)
	for r in holiday_sheet_rows(LIST_NAME, rows):
		ws.append(r)

	ws2 = wb.create_sheet("Optional Only")
	ws2.append(HOLIDAY_HEADERS)
	for r in holiday_sheet_rows(OPTIONAL_LIST_NAME, [(d, l, 0) for d, l in OPTIONAL]):
		ws2.append(r)

	wb.save(path)

	# CSV too, matching the template ERPNext handed back, so the file can be
	# uploaded without a spreadsheet round-trip changing a date format.
	import csv as _csv
	csv_path = path.rsplit(".", 1)[0] + ".csv"
	with open(csv_path, "w", newline="", encoding="utf-8-sig") as fh:
		w = _csv.writer(fh)
		w.writerow(HOLIDAY_HEADERS)
		w.writerows(holiday_sheet_rows(LIST_NAME, rows))

	return rows


def write_designations(path, names):
	wb = Workbook()
	ws = wb.active
	ws.title = "Designation"
	ws.append(["ID", "Designation"])
	for n in names:
		ws.append([n, n])
	wb.save(path)


def read_needed_designations():
	import csv
	import io
	path = os.path.join(OUT_DIR, "masters-needed.csv")
	if not os.path.exists(path):
		sys.exit("Run tools/build_employee_import.py first - " + path + " is missing")
	out = set()
	with io.open(path, encoding="utf-8-sig") as fh:
		for row in csv.DictReader(fh):
			if row["Doctype"] == "Designation" and row["Value"]:
				out.add(row["Value"])
	return sorted(out)


def main():
	os.makedirs(OUT_DIR, exist_ok=True)

	hpath = os.path.join(OUT_DIR, "01-holiday-list.xlsx")
	rows = write_holiday_list(hpath)
	named = [r for r in rows if not r[2]]
	print("wrote {0}".format(hpath))
	print("   holiday year      {0} to {1}".format(YEAR_START, YEAR_END))
	print("   total rows        {0}".format(len(rows)))
	print("   Sundays           {0}".format(len(rows) - len(named)))
	print("   named holidays    {0}".format(len(named)))
	print()
	for d, label, _ in named:
		print("      {0}  {1:<3} {2}".format(d.isoformat(), d.strftime("%a"), label))

	clashes = [(d, l) for d, l in
	           [(d, l) for d, l, w in rows if not w] if d.weekday() == WEEKLY_OFF_WEEKDAY]
	if clashes:
		print("\n   NOTE - these fall on a Sunday and are listed once, as the holiday:")
		for d, l in clashes:
			print("      {0}  {1}".format(d.isoformat(), l))

	dpath = os.path.join(OUT_DIR, "02-designation.xlsx")
	names = read_needed_designations()
	write_designations(dpath, names)
	print("\nwrote {0}  ({1} designations)".format(dpath, len(names)))


if __name__ == "__main__":
	main()
