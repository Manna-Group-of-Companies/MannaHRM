"""What the punch history says about each named shift.

    python tools/infer_shifts.py

Joins three years of real punches to the shift each person is named against in
the Factor HR master, and reports what those people actually do. The output is
evidence to confirm, **not** a shift definition.

## Why this is not the same as knowing the shift

The gap between when a shift starts and when people punch **is the
late-coming rule**. MT-003's shift is known to start at 09:30 and his median
punch is 09:29 — but on the days he arrives at 09:36, Factor HR records
`Late Coming By 00:06`. Derive the start time from his punches and you would
get 09:30 by luck; derive it from somebody habitually late and you would write
their lateness into the definition and never flag it again.

So every figure here is labelled as observed. What it is good for is narrowing
20 blank rows into 20 questions with a number already in them, and for the two
things behaviour genuinely does reveal:

  - **Whether a shift crosses midnight.** A day whose OUT precedes its IN is
    not an anomaly, it is a night shift, and no amount of policy tells you that.
  - **Whether a shift is one pattern or several.** A "24hr shift" with two
    clear clusters of start time is a rota, and configures completely
    differently from a single long window.
"""

import json
import os
import statistics
import sys
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta

try:
	from openpyxl import load_workbook
except ImportError:
	sys.exit("Needs openpyxl")

PUNCHES = "data/out/device-punches-MRP-GATE1.json"
MASTER = "data/factohr/Employee Detail Report.xlsx"

BASE = os.environ.get("ERP_URL", "https://mannarubber.m.frappe.cloud")
AUTH = "token {0}:{1}".format(os.environ.get("ERP_KEY", ""), os.environ.get("ERP_SECRET", ""))

# A shift is attributed to the day it STARTED. Anything before this hour is
# treated as the tail of the night before rather than a very early start —
# 03:08 is somebody going home, not somebody arriving.
NIGHT_CUTOFF_HOUR = 4


def erp_employees():
	out, start = [], 0
	while True:
		q = urllib.parse.urlencode({
			"fields": json.dumps(["employee_name", "employee_number", "attendance_device_id", "company"]),
			"limit_page_length": 100, "limit_start": start})
		r = urllib.request.Request(BASE + "/api/resource/Employee?" + q,
			headers={"Authorization": AUTH, "Accept": "application/json"})
		with urllib.request.urlopen(r, timeout=90) as x:
			page = json.loads(x.read().decode())["data"]
		out.extend(page)
		if len(page) < 100:
			return out
		start += 100


def shift_by_code():
	wb = load_workbook(MASTER, read_only=True, data_only=True)
	rows = [list(r) for r in wb.worksheets[0].iter_rows(max_row=3000, values_only=True)]
	wb.close()
	hi = next(i for i, r in enumerate(rows[:15])
	          if any(c and str(c).strip().lower() == "emp code" for c in r))
	H = [str(c).strip().lower() if c is not None else "" for c in rows[hi]]
	ic, ish = H.index("emp code"), H.index("working shift")
	out = {}
	for r in rows[hi + 1:]:
		if ic < len(r) and r[ic]:
			out[str(r[ic]).strip()] = (str(r[ish]).strip() if ish < len(r) and r[ish] else "")
	return out


def hhmm(minutes):
	minutes = int(round(minutes)) % (24 * 60)
	return "{0:02d}:{1:02d}".format(minutes // 60, minutes % 60)


def main():
	if not os.environ.get("ERP_KEY"):
		sys.exit("Set ERP_URL, ERP_KEY, ERP_SECRET first.")

	punches = json.load(open(PUNCHES))
	by_code = shift_by_code()
	emps = erp_employees()
	dev_to_code = {str(e["attendance_device_id"]).strip(): e["employee_number"]
	               for e in emps if e.get("attendance_device_id") and e.get("employee_number")}

	# device user -> shift name, via employee code
	dev_shift = {}
	for dev, code in dev_to_code.items():
		s = by_code.get(code)
		if s:
			dev_shift[dev] = s

	# Group each person's punches into working days. A punch before the night
	# cutoff belongs to the previous day's shift.
	per_shift_in = defaultdict(list)
	per_shift_out = defaultdict(list)
	per_shift_span = defaultdict(list)
	per_shift_people = defaultdict(set)
	crossers = defaultdict(int)
	days_seen = defaultdict(int)

	byuserday = defaultdict(list)
	for p in punches:
		u = p["u"]
		if u not in dev_shift:
			continue
		t = datetime.strptime(p["t"], "%Y-%m-%d %H:%M:%S")
		day = (t - timedelta(days=1)).date() if t.hour < NIGHT_CUTOFF_HOUR else t.date()
		byuserday[(u, day)].append((t, p["p"]))

	for (u, day), rows in byuserday.items():
		rows.sort()
		shift = dev_shift[u]
		ins = [t for t, code in rows if code == 0]
		outs = [t for t, code in rows if code == 1]
		if not ins or not outs:
			continue
		first_in, last_out = min(ins), max(outs)
		if last_out <= first_in:
			continue
		span = (last_out - first_in).total_seconds() / 3600.0
		if span > 20:
			continue  # a missed punch-out paired with the next day's punch-in

		per_shift_people[shift].add(u)
		days_seen[shift] += 1
		per_shift_in[shift].append(first_in.hour * 60 + first_in.minute)
		per_shift_out[shift].append(last_out.hour * 60 + last_out.minute)
		per_shift_span[shift].append(span)
		if last_out.date() != first_in.date():
			crossers[shift] += 1

	print("OBSERVED PUNCH BEHAVIOUR, BY THE SHIFT PEOPLE ARE NAMED AGAINST")
	print("source: {0} punches from BIO-MRP-GATE1, joined via attendance_device_id".format(len(punches)))
	print("=" * 108)
	print("{0:44} {1:>4} {2:>6} {3:>13} {4:>13} {5:>7} {6:>8}".format(
		"Shift (Factor HR name)", "ppl", "days", "typical IN", "typical OUT", "hours", "crosses"))
	print("-" * 108)

	for shift in sorted(per_shift_in, key=lambda s: -days_seen[s]):
		ins, outs = per_shift_in[shift], per_shift_out[shift]
		spans = per_shift_span[shift]
		# Median, not mean: one person who forgot to punch out until midnight
		# should not drag a whole shift's apparent end time.
		mi, mo = statistics.median(ins), statistics.median(outs)
		spread = "{0}-{1}".format(hhmm(statistics.quantiles(ins, n=10)[0]),
		                          hhmm(statistics.quantiles(ins, n=10)[8])) if len(ins) > 10 else ""
		cross = crossers[shift]
		print("{0:44} {1:>4} {2:>6} {3:>13} {4:>13} {5:>7.1f} {6:>8}".format(
			shift[:44], len(per_shift_people[shift]), days_seen[shift],
			hhmm(mi), hhmm(mo), statistics.median(spans),
			"{0}%".format(round(100 * cross / max(1, days_seen[shift]))) if cross else "-"))
		if spread:
			print("{0:44} {1}".format("", "middle 80% of starts: " + spread))

	missing = sorted(set(by_code.values()) - set(per_shift_in) - {""})
	print("\nNO PUNCH DATA HERE FOR {0} SHIFT(S) — their people punch at other gates:".format(len(missing)))
	for s in missing:
		print("   " + s)


if __name__ == "__main__":
	main()
