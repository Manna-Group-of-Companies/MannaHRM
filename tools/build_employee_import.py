"""Turn the Factor HR employee master into an ERPNext Data Import file.

    python tools/build_employee_import.py                 # active staff only
    python tools/build_employee_import.py --include-left  # everyone, leavers too

Writes `data/out/employee-import.csv`, which is gitignored along with everything
else under `data/`.

Two things this deliberately does **not** do:

  - **It never guesses a company.** Six spellings in Factor HR map onto five
    ERPNext companies through the table below, and a row whose company is not in
    that table is rejected rather than defaulted. A guessed company is a person
    in the wrong payroll, discovered at month end.
  - **It does not set `reports_to`.** A manager has to exist before anyone can
    report to them, so that is a second pass — see `--pass2`.
"""

import argparse
import csv
import os
import re
import sys
from collections import Counter

try:
	from openpyxl import load_workbook
except ImportError:
	sys.exit("Needs openpyxl: python -m pip install openpyxl")

SOURCE = "data/factohr/Employee Detail Report.xlsx"
OUT_DIR = "data/out"

# Factor HR spelling -> ERPNext Company. Confirmed with Manna on 22 Aug 2026.
#
# Note the two rows that collapse onto Manna Treads: HI-TECH PRETREADS is the
# same company, not a separate one, so its 112 people join the 4 already there.
COMPANY = {
	"MANNA RUBBER PRODUCTS PVT.LTD.": "Manna Rubber Products Private Limited",
	"HI-TECH PRETREADS": "Manna Treads",
	"MANNA TREADS PVT.LTD": "Manna Treads",
	"HI-TECH RUBBER INDUSTRIES": "Hi-Tech Rubber Industries",
	"MANNA TYRE RETREADS": "Manna Tyre Retreads",
	"MANNA GROUP H-QTRS": "Manna Group Headquarters",
}

# ERPNext names a Department "<name> - <company abbr>", so a bare "Production"
# links to nothing and the import fails a row at a time. Designation is not
# suffixed. This caught us on the first build.
ABBR = {
	"Manna Rubber Products Private Limited": "MRPPL",
	"Manna Treads": "MT",
	"Manna Tyre Retreads": "MTR",
	"Hi-Tech Rubber Industries": "HRI",
	"Manna Group Headquarters": "MGHQ",
	"Manna Tyre UAE": "MRU",
}

# ERPNext's Employee.status is a Select with exactly these options.
STATUS = {"active": "Active", "inactive": "Left"}

GENDER = {"male": "Male", "female": "Female"}

# Columns written, in the order ERPNext's importer likes to see them.
HEADERS = [
	"ID",
	"Employee Name",
	"First Name",
	"Last Name",
	"Gender",
	"Date of Birth",
	"Date of Joining",
	"Relieving Date",
	"Status",
	"Company",
	"Department",
	"Designation",
	"Employment Type",
	"Attendance Device ID (Biometric/RF tag ID)",
	"Cell Number",
	"Factor HR ID",
]


def find_header(rows, want="Emp Code", limit=15):
	target = want.lower()
	for i, row in enumerate(rows[:limit]):
		for cell in row:
			if cell and str(cell).strip().lower() == target:
				return i
	return None


def col(headers, *names):
	lowered = [h.lower() for h in headers]
	for name in names:
		if name.lower() in lowered:
			return lowered.index(name.lower())
	for name in names:
		n = name.lower()
		for i, h in enumerate(lowered):
			if h.startswith(n):
				return i
	return None


def get(row, i):
	if i is None or i >= len(row):
		return ""
	v = row[i]
	return "" if v is None else str(v).strip()


def as_date(value):
	"""Factor HR emits '1966-12-04 00:00:00'. ERPNext wants '1966-12-04'."""
	if not value:
		return ""
	m = re.match(r"(\d{4}-\d{2}-\d{2})", value)
	if m:
		return m.group(1)
	m = re.match(r"(\d{2})-([A-Za-z]{3})-(\d{4})", value)
	if m:
		months = dict(jan="01", feb="02", mar="03", apr="04", may="05", jun="06",
		              jul="07", aug="08", sep="09", oct="10", nov="11", dec="12")
		mon = months.get(m.group(2).lower())
		if mon:
			return "{0}-{1}-{2}".format(m.group(3), mon, m.group(1))
	return ""


def split_name(full):
	"""ERPNext requires `first_name`; `last_name` is optional.

	Many rows here are a single word, and several are a full name in one field
	with no reliable split. Putting everything in `first_name` and letting
	`employee_name` carry the display form is honest — inventing a surname from
	the last token would corrupt names that do not work that way.
	"""
	full = " ".join(full.split())
	if not full:
		return "", ""
	parts = full.split(" ")
	if len(parts) == 1:
		return parts[0], ""
	return " ".join(parts[:-1]), parts[-1]


def tidy(value):
	"""Factor HR shouts everything. Title-case it, but leave initialisms alone."""
	value = " ".join(value.split())
	if not value:
		return ""
	out = []
	for word in value.split(" "):
		# Q.C., H.R, PVT — anything with a dot or 3 letters or fewer that is all
		# caps is almost certainly an initialism, not a shouted word.
		if "." in word or (word.isupper() and len(word) <= 3):
			out.append(word)
		else:
			out.append(word.capitalize())
	return " ".join(out)


def department_name(raw, company):
	if not raw:
		return ""
	return "{0} - {1}".format(tidy(raw), ABBR[company])


def main():
	ap = argparse.ArgumentParser()
	ap.add_argument("--source", default=SOURCE)
	ap.add_argument("--include-left", action="store_true",
	                help="also export the 344 leavers, as Status=Left")
	args = ap.parse_args()

	book = load_workbook(args.source, read_only=True, data_only=True)
	ws = book.worksheets[0]
	rows = [list(r) for r in ws.iter_rows(max_row=3000, values_only=True)]
	book.close()

	index = find_header(rows)
	if index is None:
		sys.exit("Could not find an 'Emp Code' column in " + args.source)
	headers = [str(c).strip() if c is not None else "" for c in rows[index]]

	ix = {
		"code": col(headers, "Emp Code"),
		"name": col(headers, "Full Name"),
		"company": col(headers, "Company Name"),
		"status": col(headers, "Status"),
		"machine": col(headers, "Machine Code"),
		"dob": col(headers, "Birth Date"),
		"doj": col(headers, "Joining Date"),
		"left": col(headers, "Leaving Date"),
		"gender": col(headers, "Gender"),
		"dept": col(headers, "Department"),
		"desig": col(headers, "Designation"),
		"etype": col(headers, "Employment Type"),
		"mobile": col(headers, "Mobile No"),
	}
	missing = [k for k, v in ix.items() if v is None]
	if missing:
		sys.exit("Columns not found in the export: " + ", ".join(missing))

	from collections import defaultdict
	needed = defaultdict(set)
	out_rows, rejected = [], []
	seen_codes = set()
	stats = Counter()

	for row in rows[index + 1:]:
		code = get(row, ix["code"])
		if not code:
			continue

		status_raw = get(row, ix["status"]).lower()
		status = STATUS.get(status_raw)
		if not status:
			rejected.append((code, "unrecognised status '{0}'".format(status_raw)))
			continue

		if status == "Left" and not args.include_left:
			stats["skipped_leaver"] += 1
			continue

		if code in seen_codes:
			rejected.append((code, "duplicate employee code"))
			continue
		seen_codes.add(code)

		company = COMPANY.get(get(row, ix["company"]).upper())
		if not company:
			rejected.append((code, "company not in the mapping: '{0}'".format(get(row, ix["company"]))))
			continue

		doj = as_date(get(row, ix["doj"]))
		if not doj:
			# ERPNext refuses an Employee with no joining date, and a guessed one
			# would quietly change somebody's length of service and gratuity.
			rejected.append((code, "no usable joining date"))
			continue

		full = get(row, ix["name"])
		first, last = split_name(full)
		gender = GENDER.get(get(row, ix["gender"]).lower(), "")
		if not gender:
			stats["blank_gender"] += 1

		device = get(row, ix["machine"])
		if status == "Active" and not device:
			stats["active_without_device_id"] += 1

		out_rows.append({
			"ID": "",
			"Employee Name": full,
			"First Name": first,
			"Last Name": last,
			"Gender": gender,
			"Date of Birth": as_date(get(row, ix["dob"])),
			"Date of Joining": doj,
			"Relieving Date": as_date(get(row, ix["left"])),
			"Status": status,
			"Company": company,
			"Department": department_name(get(row, ix["dept"]), company),
			"Designation": tidy(get(row, ix["desig"])),
			"Employment Type": get(row, ix["etype"]),
			"Attendance Device ID (Biometric/RF tag ID)": device,
			"Cell Number": get(row, ix["mobile"]),
			# No Holiday List column on purpose. It is set once on each Company
			# as `default_holiday_list`, and Employee falls back to it. Repeating
			# the name on 160 rows means 160 rows to edit next April.
			"Factor HR ID": code,
		})
		stats["exported"] += 1
		stats["company:" + company] += 1
		needed["Department"].add(department_name(get(row, ix["dept"]), company))
		needed["Designation"].add(tidy(get(row, ix["desig"])))
		needed["Employment Type"].add(get(row, ix["etype"]))

	os.makedirs(OUT_DIR, exist_ok=True)
	out_path = os.path.join(OUT_DIR, "employee-import.csv")
	with open(out_path, "w", newline="", encoding="utf-8-sig") as fh:
		w = csv.DictWriter(fh, fieldnames=HEADERS)
		w.writeheader()
		w.writerows(out_rows)

	try:
		from openpyxl import Workbook
		wb = Workbook(); ws = wb.active; ws.title = "Employee"
		ws.append(HEADERS)
		for r in out_rows:
			ws.append([r[h] for h in HEADERS])
		xlsx_path = os.path.join(OUT_DIR, "03-employee.xlsx")
		wb.save(xlsx_path)
		print("wrote {0}".format(xlsx_path))
	except ImportError:
		pass

	print("wrote {0}  ({1} rows)".format(out_path, len(out_rows)))
	print()
	for key in sorted(stats):
		print("  {0:44} {1}".format(key, stats[key]))

	# Every Link value the import will need. A missing one fails that row at
	# import time, one row at a time, which is a slow way to find out.
	masters_path = os.path.join(OUT_DIR, "masters-needed.csv")
	with open(masters_path, "w", newline="", encoding="utf-8-sig") as fh:
		w = csv.writer(fh)
		w.writerow(["Doctype", "Value"])
		for dt in sorted(needed):
			for value in sorted(v for v in needed[dt] if v):
				w.writerow([dt, value])
	print("")
	print("masters the import will need -> {0}".format(masters_path))
	for dt in sorted(needed):
		print("  {0:18} {1} distinct".format(dt, len([v for v in needed[dt] if v])))

	if rejected:
		rej_path = os.path.join(OUT_DIR, "employee-rejected.csv")
		with open(rej_path, "w", newline="", encoding="utf-8-sig") as fh:
			w = csv.writer(fh)
			w.writerow(["Emp Code", "Why"])
			w.writerows(rejected)
		print("\n  *** {0} row(s) rejected -> {1}".format(len(rejected), rej_path))
		for code, why in rejected[:10]:
			print("      {0:12} {1}".format(code, why))


if __name__ == "__main__":
	main()
