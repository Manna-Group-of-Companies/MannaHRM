"""Backfill `employee_number` and `reports_to` on the imported employees.

    python tools/backfill_employee_links.py            # dry run, writes nothing
    python tools/backfill_employee_links.py --apply

Two passes, and the order is forced: `reports_to` is expressed in Factor HR as
an employee *code* ("HPT-072 - AJITH S"), so the codes have to be on the records
before anybody can be pointed at their manager.

## Why `employee_number` and not a custom field

The Factor HR code is exactly what `employee_number` is for — the employer's own
identifier for a person. A custom field would mean the same fact in two places,
and `employee_number` is already indexed, already on the standard form, and
already what every future import will match on.

## Matching, and why it refuses rather than guesses

ERPNext assigns its own names (`HR-EMP-00042`), so the two systems share no key.
Rows are matched on **name + company + joining date** together. Any employee
matching zero or more than one Factor HR row is reported and skipped — never
resolved by picking the first. A wrong match here writes one person's code onto
another, and everything downstream inherits it silently.
"""

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict

try:
	from openpyxl import load_workbook
except ImportError:
	sys.exit("Needs openpyxl: python -m pip install openpyxl")

SOURCE = "data/factohr/Employee Detail Report.xlsx"
BASE = os.environ.get("ERP_URL", "https://mannarubber.m.frappe.cloud")
AUTH = "token {0}:{1}".format(os.environ.get("ERP_KEY", ""), os.environ.get("ERP_SECRET", ""))


def req(path, method="GET", payload=None):
	data = json.dumps(payload).encode() if payload is not None else None
	r = urllib.request.Request(BASE + path, data=data, method=method,
		headers={"Authorization": AUTH, "Accept": "application/json",
		         "Content-Type": "application/json"})
	try:
		with urllib.request.urlopen(r, timeout=120) as x:
			return json.loads(x.read().decode())
	except urllib.error.HTTPError as e:
		raise SystemExit("HTTP {0} on {1}: {2}".format(e.code, path[:70], e.read().decode()[:250]))


def fetch_employees():
	out, start = [], 0
	while True:
		q = urllib.parse.urlencode({
			"fields": json.dumps(["name", "employee_name", "company", "date_of_joining",
			                      "employee_number", "reports_to", "status"]),
			"limit_page_length": 100, "limit_start": start, "order_by": "creation asc"})
		page = req("/api/resource/Employee?" + q)["data"]
		out.extend(page)
		if len(page) < 100:
			return out
		start += 100


# ------------------------------------------------------------------ source ---

COMPANY = {
	"MANNA RUBBER PRODUCTS PVT.LTD.": "Manna Rubber Products Private Limited",
	"HI-TECH PRETREADS": "Manna Treads",
	"MANNA TREADS PVT.LTD": "Manna Treads",
	"HI-TECH RUBBER INDUSTRIES": "Hi-Tech Rubber Industries",
	"MANNA TYRE RETREADS": "Manna Tyre Retreads",
	"MANNA GROUP H-QTRS": "Manna Group Headquarters",
}


def cell(row, i):
	if i is None or i >= len(row):
		return ""
	v = row[i]
	return "" if v is None else str(v).strip()


def as_date(v):
	m = re.match(r"(\d{4}-\d{2}-\d{2})", v or "")
	return m.group(1) if m else ""


def read_source():
	book = load_workbook(SOURCE, read_only=True, data_only=True)
	ws = book.worksheets[0]
	rows = [list(r) for r in ws.iter_rows(max_row=3000, values_only=True)]
	book.close()

	head = None
	for i, r in enumerate(rows[:15]):
		if any(c and str(c).strip().lower() == "emp code" for c in r):
			head = i
			break
	if head is None:
		sys.exit("no 'Emp Code' column in " + SOURCE)

	headers = [str(c).strip().lower() if c is not None else "" for c in rows[head]]

	def col(*names):
		for n in names:
			if n.lower() in headers:
				return headers.index(n.lower())
		return None

	ix = {"code": col("emp code"), "name": col("full name"), "company": col("company name"),
	      "doj": col("joining date"), "mgr": col("reporting manager"), "status": col("status")}

	out = []
	for r in rows[head + 1:]:
		code = cell(r, ix["code"])
		if not code:
			continue
		out.append({
			"code": code,
			"name": " ".join(cell(r, ix["name"]).split()),
			"company": COMPANY.get(cell(r, ix["company"]).upper(), ""),
			"doj": as_date(cell(r, ix["doj"])),
			"manager_raw": cell(r, ix["mgr"]),
			"status": cell(r, ix["status"]).lower(),
		})
	return out


def manager_code(raw):
	"""'HPT-072 - AJITH S' -> 'HPT-072'. Blank when there is no manager."""
	if not raw:
		return ""
	return raw.split(" - ", 1)[0].strip()


# ------------------------------------------------------------------- match ---

def key(name, company, doj):
	return (" ".join((name or "").upper().split()), company or "", doj or "")


def main():
	ap = argparse.ArgumentParser()
	ap.add_argument("--apply", action="store_true", help="actually write; otherwise dry run")
	args = ap.parse_args()

	if not os.environ.get("ERP_KEY"):
		sys.exit("Set ERP_URL, ERP_KEY and ERP_SECRET in the environment first.")

	source = read_source()
	employees = fetch_employees()
	print("factor hr rows {0}   erpnext employees {1}\n".format(len(source), len(employees)))

	by_key = defaultdict(list)
	for s in source:
		by_key[key(s["name"], s["company"], s["doj"])].append(s)

	matched, unmatched, ambiguous = {}, [], []
	for e in employees:
		k = key(e["employee_name"], e["company"], str(e["date_of_joining"] or "")[:10])
		hits = by_key.get(k, [])
		if len(hits) == 1:
			matched[e["name"]] = (e, hits[0])
		elif not hits:
			unmatched.append(e)
		else:
			ambiguous.append((e, hits))

	print("PASS 1 — employee_number")
	print("   matched     {0}".format(len(matched)))
	print("   unmatched   {0}".format(len(unmatched)))
	print("   ambiguous   {0}".format(len(ambiguous)))
	for e in unmatched[:10]:
		print("      no source row: {0}  {1}  {2}".format(e["name"], e["employee_name"][:24], e["company"][:22]))
	for e, hits in ambiguous[:10]:
		print("      {0} matches {1} rows: {2}".format(e["employee_name"][:24], len(hits),
		                                               ", ".join(h["code"] for h in hits)))

	wrote = 0
	if args.apply:
		for erp_name, (e, s) in matched.items():
			if e.get("employee_number") == s["code"]:
				continue
			req("/api/resource/Employee/" + urllib.parse.quote(erp_name), "PUT",
			    {"employee_number": s["code"]})
			wrote += 1
		print("   written     {0}".format(wrote))
	else:
		print("   would write {0}".format(sum(1 for _, (e, s) in matched.items()
		                                      if e.get("employee_number") != s["code"])))

	# ---- pass 2 ----
	code_to_erp = {s["code"]: erp for erp, (e, s) in matched.items()}

	print("\nPASS 2 — reports_to")
	plan, no_mgr, missing_mgr, self_ref = [], 0, [], 0
	for erp_name, (e, s) in matched.items():
		mcode = manager_code(s["manager_raw"])
		if not mcode:
			no_mgr += 1
			continue
		target = code_to_erp.get(mcode)
		if not target:
			missing_mgr.append((s["code"], mcode))
			continue
		if target == erp_name:
			# Factor HR lets somebody be their own reporting manager. ERPNext
			# rejects it, and it would make an approval chain that never ends.
			self_ref += 1
			continue
		plan.append((erp_name, target))

	print("   to set             {0}".format(len(plan)))
	print("   no manager named   {0}".format(no_mgr))
	print("   manager not loaded {0}".format(len(missing_mgr)))
	print("   self-reporting     {0}  (skipped)".format(self_ref))
	for code, mcode in missing_mgr[:10]:
		print("      {0} reports to {1}, which is not among the loaded employees".format(code, mcode))

	if args.apply:
		done = 0
		for erp_name, target in plan:
			req("/api/resource/Employee/" + urllib.parse.quote(erp_name), "PUT",
			    {"reports_to": target})
			done += 1
		print("   written            {0}".format(done))


if __name__ == "__main__":
	main()
