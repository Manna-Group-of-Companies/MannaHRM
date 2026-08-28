"""Produce a fill-in sheet for the 23 shifts, so Manna states only what we cannot read.

    python tools/build_shift_template.py   ->  data/out/04-shifts-TO-FILL.xlsx

The shift names come from the Factor HR employee master, with live headcounts.
Two timings were recovered from the attendance reports, where Factor HR appends
the window to the shift name for display; those are pre-filled and marked. The
other 21 exist nowhere in the exports and only Manna knows them.

Every column here is something the auto-attendance engine cannot work without,
and guessing any of them mis-states somebody's day:

  - **Start / End** decide which punches belong to which shift at all.
  - **Crosses midnight** decides which *day* a night shift's hours land on. Get
    it wrong and a night worker reads as absent two days running: no punch-out
    on the first, no punch-in on the second.
  - **Break** comes off worked hours, and worked hours decide half-day and
    absent. Factor HR tracks two kinds separately, so both are asked for.
  - **Rotating** is the question behind the 22- and 24-hour shifts. A 24-hour
    shift is not somebody working 24 hours; it is either a rota or a window
    inside which any 8 count, and those configure completely differently.
"""

import os
import sys

try:
	from openpyxl import Workbook, load_workbook
	from openpyxl.styles import Font, PatternFill, Alignment
	from openpyxl.utils import get_column_letter
except ImportError:
	sys.exit("Needs openpyxl: python -m pip install openpyxl")

SOURCE = "data/factohr/Employee Detail Report.xlsx"
OUT = "data/out/04-shifts-TO-FILL.xlsx"

# Median first-IN and last-OUT actually punched, from three years of records on
# BIO-MRP-GATE1, joined to the shift each person is named against. OBSERVED, not
# defined -- the gap between the two is the late-coming rule, so these are a
# starting point to correct, never a definition to adopt.
OBSERVED = {
	"Manna Rubber Products Pvt.Ltd-Production8hrshift1": ("08:24","20:30",12.1,6282,35),
	"Manna Rubber Products Pvt.Ltd-Production12hrshift1": ("08:26","20:31",12.1,1638,23),
	"Manna Rubber Products Pvt.Ltd-Office Shift": ("08:24","17:42",9.3,1956,5),
	"Hi-Tech Rubber Industries-Production shift1": ("08:20","20:30",12.1,726,4),
	"Hi-Tech Rubber Industries-Cook shift": ("06:40","15:01",8.4,224,1),
	"Hi-Tech Pretreads-Office shift": ("08:26","17:34",9.1,212,1),
	"Hi-Tech Pretreads-Production shift1": ("08:22","20:32",12.2,91,1),
	"Hi-Tech Rubber Industries-Production shift2": ("08:27","20:31",12.1,1,1),
}


# Recovered from the attendance reports, where the window is appended to the
# name for display: "... -Office shift (09:30-18:30)".
KNOWN = {
	# Stated outright by Factor HR, in the Shift Begin/End columns of the Daily
	# Attendance Detail report.
	"Manna Treads Pvt.Ltd-Office shift": ("09:30", "18:30", "stated by Factor HR - please confirm"),
	# Factor HR appends the window to the shift name in some reports.
	"Hi-Tech Pretreads-Other location": ("09:30", "18:30", "read from the shift name - please confirm"),
	# INFERRED, not stated. One person (MRP-004), 19 days, median first punch
	# 08:32 and median last 17:41. Calibrated against MT-003, whose shift IS
	# known: he punches a median 09:29 against a 09:30 start and 18:50 against
	# an 18:30 end. Same offsets applied here give roughly 08:30-17:30.
	"Manna Rubber Products Pvt.Ltd-Office Shift": ("08:30", "17:30", "INFERRED from one person's punches - must be confirmed, not assumed"),
}

COLUMNS = [
	("Shift (Factor HR name)", 44),
	("People", 8),
	("Observed IN", 12),
	("Observed OUT", 12),
	("Observed hrs", 12),
	("Days seen", 10),
	("Start", 10),
	("End", 10),
	("Crosses midnight?", 18),
	("Break start", 12),
	("Break end", 12),
	("Rotating?", 12),
	("Half day below (hrs)", 20),
	("Absent below (hrs)", 18),
	("Notes", 40),
]


def shift_counts():
	wb = load_workbook(SOURCE, read_only=True, data_only=True)
	ws = wb.worksheets[0]
	rows = [list(r) for r in ws.iter_rows(max_row=3000, values_only=True)]
	wb.close()

	head = None
	for i, r in enumerate(rows[:15]):
		if any(c and str(c).strip().lower() == "emp code" for c in r):
			head = i
			break
	headers = [str(c).strip().lower() if c is not None else "" for c in rows[head]]
	i_shift = headers.index("working shift")
	i_status = headers.index("status")

	counts = {}
	for r in rows[head + 1:]:
		if i_status >= len(r) or not r[i_status]:
			continue
		if str(r[i_status]).strip().lower() != "active":
			continue
		v = r[i_shift] if i_shift < len(r) else None
		if not v:
			continue
		name = str(v).strip()
		counts[name] = counts.get(name, 0) + 1
	return counts


def main():
	os.makedirs("data/out", exist_ok=True)
	counts = shift_counts()

	wb = Workbook()
	ws = wb.active
	ws.title = "Shifts"

	head_font = Font(bold=True, color="FFFFFF")
	head_fill = PatternFill("solid", fgColor="0E6B73")
	known_fill = PatternFill("solid", fgColor="E3EFE7")
	ask_fill = PatternFill("solid", fgColor="F6EDDA")

	ws.append([c for c, _ in COLUMNS])
	for i, (_, width) in enumerate(COLUMNS, start=1):
		ws.column_dimensions[get_column_letter(i)].width = width
		cell = ws.cell(row=1, column=i)
		cell.font = head_font
		cell.fill = head_fill
		cell.alignment = Alignment(wrap_text=True, vertical="center")
	ws.freeze_panes = "A2"

	for name, n in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
		hit = KNOWN.get(name)
		start, end, note = hit if hit else ("", "", "")
		# The long-window shifts are the ones that need explaining, so say so
		# on the row rather than in a covering note nobody reads.
		low = name.lower()
		if "24hr" in low or "22hr" in low:
			note = "is this a rota, or a window inside which any 8 hours count?"
		elif "12hr" in low or "12hrs" in low or "production12" in low:
			note = "does this one cross midnight?"
		o = OBSERVED.get(name)
		if o and not start:
			# Only ever offered alongside the blank columns, never poured into
			# them: a number in the Start column reads as agreed.
			note = note or "observed only - state the real start and end"
		row = [name, n,
		       o[0] if o else "", o[1] if o else "", o[2] if o else "", o[3] if o else "",
		       start, end, "", "", "", "", "", "", note]
		ws.append(row)
		r = ws.max_row
		fill = known_fill if name in KNOWN else ask_fill
		for c in range(3, 11):
			ws.cell(row=r, column=c).fill = fill

	ws2 = wb.create_sheet("How to fill this in")
	for line in [
		("What this is", True),
		("", False),
		("The 23 shifts your 160 active staff are on, taken from the Factor HR", False),
		("employee master with live headcounts.", False),
		("", False),
		("Green rows were recovered from your attendance reports - please confirm", False),
		("rather than assume. Amber rows exist nowhere in the exports.", False),
		("", False),
		("Columns, and why each matters", True),
		("", False),
		("Start / End       which punches belong to this shift at all", False),
		("Crosses midnight  which DAY the hours land on. Wrong, and a night", False),
		("                  worker reads absent two days running.", False),
		("Break             comes off worked hours, and worked hours decide", False),
		("                  half-day and absent. Factor HR tracks two kinds.", False),
		("Rotating          the question behind the 22hr and 24hr shifts", False),
		("Half day below    hours worked under this = half day", False),
		("Absent below      hours worked under this = absent", False),
		("", False),
		("If half-day and absent thresholds are the same across every shift,", False),
		("fill the first row and say so - no need to repeat it 23 times.", False),
	]:
		ws2.append([line[0]])
		if line[1]:
			ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
	ws2.column_dimensions["A"].width = 76

	wb.save(OUT)
	print("wrote {0}".format(OUT))
	print("   {0} shifts, {1} people".format(len(counts), sum(counts.values())))
	print("   pre-filled from the exports: {0}".format(len(KNOWN)))
	print("   needing Manna to state them: {0}".format(len(counts) - len(KNOWN)))


if __name__ == "__main__":
	main()
