"""Profile a Factor HR export without printing its contents.

Reports structure — sheets, where the real header row is, column names, row
counts — and masks every sample value. That distinction is the point: knowing a
file has an `Aadhaar` column is what tells us how to map it, and the numbers
themselves are nobody's business until the load actually runs.

    python tools/inspect_export.py data/factohr/
    python tools/inspect_export.py data/factohr/employee-master.xlsx

Reads .xlsx and .csv. Needs openpyxl for .xlsx:

    python -m pip install openpyxl
"""

import csv
import os
import re
import sys

# Columns whose sample values are never shown, however harmless one row looks.
# Matched loosely on purpose — an export column called "Emp PAN No." should hit.
SENSITIVE = re.compile(
	r"aadha|pan\b|uan|account|ifsc|bank|salary|ctc|gross|net|basic|pf\b|esi|"
	r"password|dob|birth|mobile|phone|email|address|nominee",
	re.I,
)

# A header row is the first row that looks like labels rather than a title
# banner: several non-empty cells, none of them absurdly long.
MIN_HEADER_CELLS = 3
MAX_HEADER_LEN = 60


def mask(value, column):
	"""A value rendered as its shape, never its content."""
	if value is None or str(value).strip() == "":
		return "—"
	text = str(value).strip()
	if SENSITIVE.search(column or ""):
		return "<{0} chars>".format(len(text))
	if len(text) > 18:
		return text[:15] + "..."
	return text


def looks_like_header(cells):
	filled = [c for c in cells if c is not None and str(c).strip()]
	if len(filled) < MIN_HEADER_CELLS:
		return False
	return all(len(str(c)) <= MAX_HEADER_LEN for c in filled)


def find_header(rows, limit=15):
	"""Index of the real header row.

	Factor HR reports carry a title, a logo row and a date range above the
	columns. Scanning for the first row that looks like labels beats assuming
	row 0 and reading the report's own name as a column.
	"""
	for i, row in enumerate(rows[:limit]):
		if looks_like_header(row):
			return i
	return 0


def report(name, header, rows):
	print("    header row : {0}".format(header["index"] + 1))
	print("    columns    : {0}".format(len(header["columns"])))
	print("    data rows  : {0}".format(len(rows)))
	print("    ---")
	width = max((len(str(c)) for c in header["columns"]), default=10)
	width = min(max(width, 12), 38)
	for col_index, column in enumerate(header["columns"]):
		samples = []
		for row in rows[:3]:
			if col_index < len(row):
				samples.append(mask(row[col_index], column))
		flag = "  [sensitive]" if SENSITIVE.search(str(column)) else ""
		print(
			"    {0:<{1}}  {2}{3}".format(
				str(column)[:width], width, " | ".join(samples) or "—", flag
			)
		)


def inspect_xlsx(path):
	try:
		from openpyxl import load_workbook
	except ImportError:
		print("    openpyxl not installed — run: python -m pip install openpyxl")
		return

	# read_only keeps a 50k-row export from being loaded into memory whole.
	book = load_workbook(path, read_only=True, data_only=True)
	for sheet in book.worksheets:
		print("  sheet: {0}".format(sheet.title))
		rows = [list(r) for r in sheet.iter_rows(max_row=400, values_only=True)]
		if not rows:
			print("    empty")
			continue
		index = find_header(rows)
		columns = [c if c is not None else "" for c in rows[index]]
		total = sheet.max_row - index - 1 if sheet.max_row else len(rows) - index - 1
		report(
			sheet.title,
			{"index": index, "columns": columns},
			rows[index + 1 :],
		)
		if total > len(rows) - index - 1:
			print("    (first 400 rows scanned; sheet holds about {0})".format(total))
		print()
	book.close()


def inspect_csv(path):
	# Factor HR exports from an Indian tenant are usually cp1252, occasionally
	# utf-8-sig. Trying in order beats failing on one stray rupee sign.
	for encoding in ("utf-8-sig", "cp1252", "latin-1"):
		try:
			with open(path, newline="", encoding=encoding) as handle:
				rows = [r for _, r in zip(range(400), csv.reader(handle))]
			break
		except UnicodeDecodeError:
			continue
	else:
		print("    could not decode as utf-8, cp1252 or latin-1")
		return

	if not rows:
		print("    empty")
		return

	index = find_header(rows)
	print("  (csv, decoded as {0})".format(encoding))
	report(os.path.basename(path), {"index": index, "columns": rows[index]}, rows[index + 1 :])
	print()


def main():
	target = sys.argv[1] if len(sys.argv) > 1 else "data/factohr"

	if os.path.isdir(target):
		names = sorted(
			n for n in os.listdir(target) if n.lower().endswith((".xlsx", ".xls", ".csv"))
		)
		paths = [os.path.join(target, n) for n in names]
		if not paths:
			print("Nothing to inspect in {0} — drop the exports there first.".format(target))
			return
	else:
		paths = [target]

	for path in paths:
		size = os.path.getsize(path) / 1024.0
		print("=" * 72)
		print("{0}  ({1:.0f} KB)".format(os.path.basename(path), size))
		print("=" * 72)

		lower = path.lower()
		if lower.endswith(".csv"):
			inspect_csv(path)
		elif lower.endswith(".xlsx"):
			inspect_xlsx(path)
		elif lower.endswith(".xls"):
			# The old binary format. openpyxl cannot read it and pulling in
			# xlrd for one file is not worth it — Excel will re-save as .xlsx.
			print("  .xls is the old binary format. Open it and Save As .xlsx.\n")
		else:
			print("  unsupported file type\n")


if __name__ == "__main__":
	main()
